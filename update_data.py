"""
update_data.py
--------------
Baca Datamart_Risiko_Operasional.xlsx (sheet DATAMART + LAPORAN)
→ generate public/data.js untuk portal web.

Jalankan setiap kali selesai update Excel, lalu git push ke GitHub.
Vercel otomatis re-deploy dalam ~30 detik.

Usage:
    python update_data.py            # jalan normal, warning ditampilkan tapi tetap lanjut
    python update_data.py --strict   # berhenti (exit code 1) kalau ada warning data
"""

import json, sys, re, argparse
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

BASE       = Path(__file__).parent
EXCEL_FILE = BASE / "Datamart_Risiko_Operasional.xlsx"
OUT_FILE   = BASE / "public" / "data.js"

# Kolom yang isinya WAJIB numerik atau persen (selain teks bebas).
# "jen" boleh cuma salah satu dari dua nilai ini — typo di sini bikin
# filter "Hasil"/"Pembentuk" di web diam-diam gagal.
JENIS_VALID = {"Hasil", "Pembentuk"}

# Pola nilai yang dianggap "bersih": angka biasa, atau persen format Indonesia (1.234,56%)
RE_NUMERIC = re.compile(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$|^-?\d+(,\d+)?$")
RE_PERCENT = re.compile(r"^-?\d{1,3}(\.\d{3})*(,\d+)?%$|^-?\d+(,\d+)?%$")

WARNINGS = []


def warn(msg):
    WARNINGS.append(msg)
    print(f"  ⚠ {msg}")


RE_RAW_PERCENT_CANDIDATE = re.compile(r"^-?0?\.\d+$|^-?1\.0+$")  # 0.xxxx atau 1.0 (gaya US, antara 0 dan 1)
RE_FORMATTED_PERCENT = re.compile(r"^-?\d+(\.\d+)?,?\d*%$")


def looks_like_unformatted_percent(s):
    """Angka desimal titik antara -1 dan 1 (mis. 0.4344) — kandidat persen yang belum diformat."""
    return bool(RE_RAW_PERCENT_CANDIDATE.match(s))


def normalize_row_values(par_clean, period_vals):
    """
    period_vals: dict {periode: val_str_asli}.
    Mengembalikan dict baru {periode: val_str_ternormalisasi}.

    Logika: kalau baris ini punya MINIMAL SATU periode lain yang sudah
    berformat persen Indonesia ("56,77%"), maka periode lain dalam baris
    yang sama yang berupa desimal mentah 0–1 (gaya US, mis. "0.4344")
    dianggap representasi persentase yang belum diformat, dan dikonversi
    ke format persen Indonesia ("43,44%") — bukan dibulatkan sebagai angka biasa.

    Kalau baris ini TIDAK punya histori format persen sama sekali, desimal
    titik dengan banyak digit tetap dibulatkan ke 2 desimal seperti biasa
    (kemungkinan itu hasil formula numerik, bukan persentase).
    """
    has_percent_history = any(
        v is not None and RE_FORMATTED_PERCENT.match(v.strip())
        for v in period_vals.values()
    )
    out = {}
    for p, val_str in period_vals.items():
        if val_str is None:
            out[p] = None
            continue
        s = val_str.strip()
        if has_percent_history and looks_like_unformatted_percent(s):
            try:
                pct = float(s) * 100
                # format Indonesia: koma sebagai desimal, 2 angka di belakang koma, trim trailing zero tak perlu
                pct_str = f"{pct:.2f}".rstrip("0").rstrip(".")
                if "." in f"{pct:.2f}" and pct_str == str(int(pct)):
                    pct_str = str(int(pct))
                out[p] = pct_str.replace(".", ",") + "%"
            except ValueError:
                out[p] = val_str
            continue
        out[p] = normalize_value(val_str)
    return out


def normalize_value(val_str):
    """
    Kalau nilai berupa angka dengan satu titik desimal (gaya US) dan bukan
    kandidat persen mentah, bulatkan ke 2 desimal supaya rapi dan konsisten
    ditampilkan di portal — baik untuk desimal panjang hasil formula
    ("405.9992105263158" → "405.99") maupun desimal pendek yang sudah rapi
    ("41921.5" → "41921.50"). Nilai lain (termasuk format Indonesia dengan
    koma) dikembalikan apa adanya.
    """
    if val_str is None:
        return None
    s = val_str.strip()
    m = re.match(r"^-?\d+\.\d+$", s)
    if m:
        try:
            return f"{float(s):.2f}"
        except ValueError:
            return val_str
    return val_str


def check_value_format(val_str, row_label, period):
    """
    Cek apakah sebuah nilai sel terlihat seperti angka/persen yang valid.
    Tidak menolak teks lain (beberapa parameter mungkin sengaja kualitatif),
    tapi nilai yang KELIHATAN seperti angka rusak (mis. "Rp 1.000.000",
    "1.234.567" ribuan salah format dengan banyak titik, sel error formula
    "#REF!") akan di-flag.

    Catatan: angka dengan SATU titik desimal (gaya US) — baik desimalnya
    panjang seperti "405.9992105263158" (hasil formula) maupun pendek
    seperti "41921.5" atau "15.95" — TIDAK di-warning. Keduanya valid
    secara matematis; satu titik desimal satu-satunya jarang merupakan
    typo. Yang tetap di-warning hanya pola dengan BANYAK titik dalam satu
    angka (mis. "1.234.567"), yang merupakan ciri khas typo pemisah ribuan
    gaya Indonesia yang salah ketik di luar konteksnya.

    Desimal antara -1 dan 1 (mis. "0.4344") yang merupakan kandidat
    persentase belum diformat juga tidak di-warning di sini — itu
    ditangani sebagai konversi persen di normalize_row_values().
    """
    if val_str is None:
        return
    s = val_str.strip()
    if s == "":
        return
    if RE_NUMERIC.match(s) or RE_PERCENT.match(s):
        return
    if looks_like_unformatted_percent(s):
        return
    if re.match(r"^[A-Z#].*$", s) and ("#" in s or "REF" in s or "DIV" in s or "N/A" in s.upper()):
        warn(f"'{row_label}' periode {period}: sel berisi kemungkinan ERROR FORMULA Excel → \"{s}\"")
        return
    if re.search(r"rp\s*\d|\d.*\.\d{3}\.\d{3}", s, re.IGNORECASE):
        warn(f"'{row_label}' periode {period}: format mencurigakan (mirip Rupiah/ribuan salah format) → \"{s}\"")
        return
    # Satu titik desimal (gaya US), berapapun panjang desimalnya, dianggap valid — tidak di-warning.
    # Selain itu dianggap teks kualitatif yang sah, tidak di-warn.


def read_datamart(wb):
    if "DATAMART" not in wb.sheetnames:
        print("ERROR: Sheet 'DATAMART' tidak ditemukan di file Excel.")
        sys.exit(1)
    ws = wb["DATAMART"]
    row3 = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    periods, period_cols = [], {}
    for ci, val in enumerate(row3):
        s = str(val).strip() if val is not None else ""
        if re.match(r"^\d{6}$", s):
            periods.append(s)
            period_cols[s] = ci + 1
    if not periods:
        print("ERROR: Tidak ada kolom periode (format YYYYMM, contoh 202604) di baris 3 sheet DATAMART.")
        sys.exit(1)
    periods.sort()

    seen_keys = set()
    rows = []
    for er in range(5, ws.max_row + 1):
        par = ws.cell(row=er, column=3).value
        if not par or str(par).strip() == "":
            continue
        key = str(ws.cell(row=er, column=1).value or "").strip()
        jen = str(ws.cell(row=er, column=4).value or "").strip()
        par_clean = str(par).strip()

        if not key:
            warn(f"Baris {er} ('{par_clean}'): kolom KEY kosong — baris ini tidak punya ID unik.")
        elif key in seen_keys:
            warn(f"Baris {er} ('{par_clean}'): KEY '{key}' duplikat — sudah dipakai parameter lain.")
        else:
            seen_keys.add(key)

        if jen and jen not in JENIS_VALID:
            warn(f"Baris {er} ('{par_clean}'): kolom JENIS berisi '{jen}', diharapkan 'Hasil' atau 'Pembentuk' — cek typo.")

        row = {
            "key": key,
            "bag": str(ws.cell(row=er, column=2).value or "").strip(),
            "par": par_clean,
            "jen": jen,
            "lnk": str(ws.cell(row=er, column=5).value or "").strip(),
        }
        raw_vals = {}
        for p in periods:
            val = ws.cell(row=er, column=period_cols[p]).value
            val_str = None if (val is None or str(val).strip() == "") else str(val).strip()
            check_value_format(val_str, par_clean, p)
            raw_vals[p] = val_str
        normalized = normalize_row_values(par_clean, raw_vals)
        for p in periods:
            row[p] = normalized[p]
        rows.append(row)

    print(f"  DATAMART : {len(rows)} parameter, {len(periods)} periode")
    return {"periods": periods, "rows": rows}


def read_laporan(wb):
    if "LAPORAN" not in wb.sheetnames:
        print("  LAPORAN  : sheet tidak ditemukan, dilewati.")
        return []
    ws = wb["LAPORAN"]
    files = []
    for er in range(5, ws.max_row + 1):
        nama = ws.cell(row=er, column=1).value
        if not nama or str(nama).strip() == "":
            continue
        aktif = str(ws.cell(row=er, column=8).value or "Y").strip().upper()
        if aktif != "Y":
            continue
        link = str(ws.cell(row=er, column=5).value or "").strip()
        tipe = str(ws.cell(row=er, column=4).value or "pdf").strip().lower()
        if not link:
            warn(f"Baris {er} LAPORAN ('{str(nama).strip()}'): kolom LINK kosong — kartu dokumen ini tidak akan bisa dibuka.")
        if tipe != "link" and link.startswith("http"):
            warn(f"Baris {er} LAPORAN ('{str(nama).strip()}'): tipe '{tipe}' tapi link berupa URL ('{link}') — mungkin maksudnya tipe 'link'?")
        files.append({
            "nama":  str(nama).strip(),
            "desc":  str(ws.cell(row=er, column=2).value or "").strip(),
            "kat":   str(ws.cell(row=er, column=3).value or "Dokumen").strip(),
            "tipe":  tipe,
            "link":  link,
            "ukuran":str(ws.cell(row=er, column=6).value or "").strip(),
            "tgl":   str(ws.cell(row=er, column=7).value or "").strip(),
        })
    print(f"  LAPORAN  : {len(files)} dokumen aktif")
    return files


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--strict", action="store_true",
                         help="Berhenti dengan exit code 1 jika ada warning data (cocok untuk CI/GitHub Actions).")
    args = parser.parse_args()

    if not EXCEL_FILE.exists():
        print(f"ERROR: File tidak ditemukan → {EXCEL_FILE}")
        sys.exit(1)

    print(f"Membaca {EXCEL_FILE.name}...")
    wb = load_workbook(EXCEL_FILE, data_only=True)
    dm      = read_datamart(wb)
    laporan = read_laporan(wb)
    ts      = datetime.now().strftime("%Y-%m-%d %H:%M")

    payload = {
        "periods": dm["periods"],
        "rows":    dm["rows"],
        "laporan": laporan,
        "updated": ts,
        "warnings": WARNINGS,   # ikut disimpan supaya bisa ditampilkan sebagai indikator di portal
    }
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    js = (
        f"// Auto-generated by update_data.py — {ts}\n"
        f"// Jangan edit manual. Edit Excel lalu jalankan script ini.\n"
        f"window.DATAMART = {json.dumps(payload, ensure_ascii=False, indent=2)};\n"
    )
    OUT_FILE.write_text(js, encoding="utf-8")

    print(f"\n✓ Berhasil → {OUT_FILE}")
    print(f"  {len(dm['rows'])} parameter · {len(dm['periods'])} periode · {len(laporan)} dokumen")

    if WARNINGS:
        print(f"\n⚠ {len(WARNINGS)} WARNING ditemukan (lihat detail di atas).")
        if args.strict:
            print("✗ Mode --strict aktif → berhenti. Perbaiki Excel lalu jalankan ulang.")
            sys.exit(1)
        else:
            print("  Tetap lanjut (mode normal). Jalankan dengan --strict untuk memblokir deploy jika ada warning.")
    else:
        print("  Tidak ada warning data. ✓")

    print(f"\nSelanjutnya:")
    print(f"  git add public/data.js Datamart_Risiko_Operasional.xlsx")
    print(f"  git commit -m 'update data {dm['periods'][-1]}'")
    print(f"  git push")


if __name__ == "__main__":
    main()
